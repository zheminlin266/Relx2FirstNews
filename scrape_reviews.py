#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
通过 Apple iTunes RSS API 抓取陌陌、探探的 App Store 评价数据，并保存到 Excel 文件。

用法:
    python scrape_reviews.py            # 抓取真实数据
    python scrape_reviews.py --demo     # 演示模式（使用示例数据，无需网络）
"""

import sys
import time
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# App 信息配置（App Store 中国区 ID）
APPS = [
    {"name": "陌陌", "app_id": "461598127"},
    {"name": "探探", "app_id": "981975024"},
]

# Apple RSS API 模板（每页最多 50 条，支持最多 10 页）
RSS_URL = "https://itunes.apple.com/cn/rss/customerreviews/page={page}/id={app_id}/sortBy=mostRecent/json"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
}

MAX_PAGES = 10  # Apple RSS API 最多支持 10 页


def fetch_reviews(app_id, app_name):
    """抓取某个 App 的所有评价（最多 10 页）"""
    reviews = []
    for page in range(1, MAX_PAGES + 1):
        url = RSS_URL.format(page=page, app_id=app_id)
        try:
            resp = requests.get(url, headers=HEADERS, timeout=20)
            print(f"  [{app_name}] 第 {page} 页 HTTP {resp.status_code}")
            if resp.status_code != 200:
                break
            data = resp.json()
        except Exception as e:
            print(f"  [{app_name}] 第 {page} 页请求失败: {e}")
            break

        entries = data.get("feed", {}).get("entry", [])
        if not entries:
            print(f"  [{app_name}] 第 {page} 页无数据，停止翻页。")
            break

        # 第一条 entry 通常是 App 本身信息（含 im:image 字段），跳过
        page_count = 0
        for entry in entries:
            if "im:image" in entry:
                continue
            try:
                review = {
                    "App名称": app_name,
                    "评价ID": entry.get("id", {}).get("label", ""),
                    "标题": entry.get("title", {}).get("label", ""),
                    "内容": entry.get("content", {}).get("label", ""),
                    "评分": entry.get("im:rating", {}).get("label", ""),
                    "版本": entry.get("im:version", {}).get("label", ""),
                    "作者": entry.get("author", {}).get("name", {}).get("label", ""),
                    "更新时间": entry.get("updated", {}).get("label", ""),
                    "投票数": entry.get("im:voteCount", {}).get("label", ""),
                    "有用票数": entry.get("im:voteSum", {}).get("label", ""),
                }
                reviews.append(review)
                page_count += 1
            except Exception as e:
                print(f"  解析评价条目失败: {e}")

        print(f"  [{app_name}] 第 {page} 页抓取 {page_count} 条，累计 {len(reviews)} 条")
        time.sleep(0.5)  # 避免请求过快

    return reviews


def get_demo_reviews():
    """返回演示用的示例数据"""
    demo_data = [
        # 陌陌示例评价
        {"App名称": "陌陌", "评价ID": "10001", "标题": "很好用的社交软件",
         "内容": "功能非常丰富，界面清晰，认识了很多朋友，推荐大家使用！", "评分": "5",
         "版本": "9.5.0", "作者": "用户A", "更新时间": "2026-03-01T10:00:00-07:00",
         "投票数": "15", "有用票数": "12"},
        {"App名称": "陌陌", "评价ID": "10002", "标题": "广告有点多",
         "内容": "基本功能不错，但是广告太多了，影响使用体验，希望能减少广告。", "评分": "3",
         "版本": "9.5.0", "作者": "用户B", "更新时间": "2026-02-28T14:30:00-07:00",
         "投票数": "8", "有用票数": "7"},
        {"App名称": "陌陌", "评价ID": "10003", "标题": "直播功能很棒",
         "内容": "直播功能做得很好，主播质量高，互动体验好。", "评分": "5",
         "版本": "9.4.0", "作者": "用户C", "更新时间": "2026-02-25T09:15:00-07:00",
         "投票数": "20", "有用票数": "18"},
        {"App名称": "陌陌", "评价ID": "10004", "标题": "隐私保护一般",
         "内容": "社交功能还可以，但感觉隐私保护做得不够好，需要改进。", "评分": "2",
         "版本": "9.3.0", "作者": "用户D", "更新时间": "2026-02-20T16:45:00-07:00",
         "投票数": "5", "有用票数": "4"},
        {"App名称": "陌陌", "评价ID": "10005", "标题": "交友体验好",
         "内容": "附近的人功能很实用，成功认识了很多朋友，体验很好。", "评分": "4",
         "版本": "9.5.0", "作者": "用户E", "更新时间": "2026-03-05T11:00:00-07:00",
         "投票数": "30", "有用票数": "25"},
        # 探探示例评价
        {"App名称": "探探", "评价ID": "20001", "标题": "滑动匹配很有趣",
         "内容": "左滑右滑的交友方式很有意思，匹配到了很多志同道合的人！", "评分": "5",
         "版本": "6.8.0", "作者": "用户F", "更新时间": "2026-03-10T08:00:00-07:00",
         "投票数": "45", "有用票数": "40"},
        {"App名称": "探探", "评价ID": "20002", "标题": "男女比例不平衡",
         "内容": "男生太多女生太少，很难匹配成功，建议平台控制一下比例。", "评分": "2",
         "版本": "6.8.0", "作者": "用户G", "更新时间": "2026-03-08T13:30:00-07:00",
         "投票数": "60", "有用票数": "55"},
        {"App名称": "探探", "评价ID": "20003", "标题": "新功能很实用",
         "内容": "最近更新的功能很棒，可以看到更多信息了，匹配效率提高了不少。", "评分": "4",
         "版本": "6.7.0", "作者": "用户H", "更新时间": "2026-03-03T10:20:00-07:00",
         "投票数": "22", "有用票数": "20"},
        {"App名称": "探探", "评价ID": "20004", "标题": "虚假账号太多",
         "内容": "遇到很多机器人账号和虚假用户，平台应该加强审核机制。", "评分": "1",
         "版本": "6.6.0", "作者": "用户I", "更新时间": "2026-02-15T15:00:00-07:00",
         "投票数": "80", "有用票数": "75"},
        {"App名称": "探探", "评价ID": "20005", "标题": "界面设计简洁",
         "内容": "界面设计很简洁美观，操作流畅，整体体验不错。", "评分": "4",
         "版本": "6.8.0", "作者": "用户J", "更新时间": "2026-03-12T09:45:00-07:00",
         "投票数": "18", "有用票数": "16"},
    ]
    return demo_data


def save_to_excel(all_reviews, filename):
    """将评价数据保存到 Excel 文件"""
    wb = openpyxl.Workbook()
    ws_all = wb.active
    ws_all.title = "全部评价"

    col_headers = ["App名称", "评价ID", "标题", "内容", "评分", "版本", "作者", "更新时间", "投票数", "有用票数"]

    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def write_headers(ws):
        for col_idx, h in enumerate(col_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        ws.row_dimensions[1].height = 25

    write_headers(ws_all)

    # 为每个 App 创建独立工作表
    app_sheets = {}
    for app in APPS:
        ws = wb.create_sheet(title=app["name"])
        write_headers(ws)
        app_sheets[app["name"]] = ws

    # 写入数据
    row_all = 2
    row_app = {app["name"]: 2 for app in APPS}

    # 交替行颜色
    fill_light = PatternFill(start_color="EEF3FB", end_color="EEF3FB", fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for i, review in enumerate(all_reviews):
        app_name = review["App名称"]
        row_data = [review.get(h, "") for h in col_headers]
        row_fill = fill_light if i % 2 == 0 else fill_white

        # 写入全部评价表
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_all.cell(row=row_all, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = row_fill
        row_all += 1

        # 写入各 App 独立表
        if app_name in app_sheets:
            ws = app_sheets[app_name]
            r = row_app[app_name]
            app_fill = fill_light if (r % 2 == 0) else fill_white
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=col_idx, value=val)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.fill = app_fill
            row_app[app_name] += 1

    # 调整列宽
    col_widths = [10, 15, 30, 60, 8, 10, 20, 25, 10, 10]
    for ws in [ws_all] + list(app_sheets.values()):
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
        # 冻结首行
        ws.freeze_panes = "A2"
        # 自动筛选
        ws.auto_filter.ref = ws.dimensions

    wb.save(filename)
    print(f"\n数据已保存至: {filename}")


def main():
    demo_mode = "--demo" in sys.argv

    if demo_mode:
        print("=== 演示模式：使用示例数据 ===")
        all_reviews = get_demo_reviews()
        for app in APPS:
            count = sum(1 for r in all_reviews if r["App名称"] == app["name"])
            print(f"[{app['name']}] 示例数据 {count} 条")
    else:
        all_reviews = []
        for app in APPS:
            print(f"\n正在抓取 [{app['name']}] (ID: {app['app_id']}) 的评价...")
            reviews = fetch_reviews(app["app_id"], app["name"])
            print(f"[{app['name']}] 共抓取 {len(reviews)} 条评价")
            all_reviews.extend(reviews)

    if not all_reviews:
        print("未抓取到任何评价数据。")
        return

    print(f"\n共评价总计: {len(all_reviews)} 条")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    mode_tag = "_demo" if demo_mode else ""
    filename = f"/home/user/Relx2FirstNews/app_reviews{mode_tag}_{timestamp}.xlsx"
    save_to_excel(all_reviews, filename)
    print(f"完成！共 {len(all_reviews)} 条评价，包含 {len(APPS)} 个 App。")


if __name__ == "__main__":
    main()
